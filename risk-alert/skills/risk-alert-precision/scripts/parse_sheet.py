#!/usr/bin/env python3
"""
parse_sheet.py
==============
Parses the Risk Intel Alerts XLSX and outputs:
  1. sheet_data.csv     — rows scoped to the last N days (Added At filter),
                          one row per payment: id, alert_codes, added_at,
                          date_reviewed, reviewed, risk_status_enum
  2. alert_links.json   — { alert_name: redash_url } (hyperlinks from Alerts cells)

Known column layout (0-indexed):
  0  Payment ID
  1  Amount
  2  Payment Status
  3  Risk Status          ← infra payment status (NOT used for fraud tagging)
  4  Created At (Payment)
  5  Org ID
  6  Alert Count
  7  Alerts               ← comma-separated alert names; cells carry Redash hyperlinks
  8  Added At             ← scope filter applied here
  9  Analyst
  10 Reviewed             ← Yes / No / empty; exclusion filter applied in calculate_precision.py
  11 Risk Status          ← ENUM: Good Payment | Bad Payment | Escalated | Off-rail
                             (used for fraud tagging — different from col 3)
  12 Fraud type
  13 Fraud Ring
  14 Notes
  15 Date reviewed

Usage:
    # Verify columns on first run against a new sheet version
    python parse_sheet.py alerts.xlsx --print-columns

    # Standard run (last 30 days)
    python parse_sheet.py alerts.xlsx --out-dir /tmp/risk_intel/

    # Post-MVP: extend scope to 90 days
    python parse_sheet.py alerts.xlsx --out-dir /tmp/risk_intel/ --scope-days 90
"""

import argparse
import csv
import json
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path


# ── Expected column positions ─────────────────────────────────────────────────
# Using positional access avoids issues with the two duplicate "Risk Status" headers.
COL = {
    'payment_id':       0,
    'alerts':           7,   # Alerts (with Redash hyperlinks)
    'added_at':         8,   # Added At — scope filter applied here
    'reviewed':         10,  # Reviewed (Yes/No)
    'risk_status_enum': 11,  # Second Risk Status: Good Payment | Bad Payment | Escalated | Off-rail
    'date_reviewed':    15,  # Date reviewed
}

# Expected headers for validation (printed during --print-columns)
EXPECTED_HEADERS = [
    'Payment ID', 'Amount', 'Payment Status', 'Risk Status',
    'Created At (Payment)', 'Org ID', 'Alert Count', 'Alerts',
    'Added At', 'Analyst', 'Reviewed', 'Risk Status',
    'Fraud type', 'Fraud Ring', 'Notes', 'Date reviewed',
]


def _cell_str(cell) -> str:
    """Return cell value as stripped string, or empty string if null."""
    return str(cell.value).strip() if cell.value is not None else ''


def parse(xlsx_path: str, out_dir: str, scope_days: int = 30,
          print_columns: bool = False):
    try:
        import openpyxl
    except ImportError:
        sys.exit("ERROR: openpyxl not installed.\nRun: pip install openpyxl --break-system-packages")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Read header row
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    actual_headers = [c.value for c in header_row]

    # ── Column verification / print mode ─────────────────────────────────────
    if print_columns:
        print("\nActual column headers (0-indexed):")
        for i, h in enumerate(actual_headers):
            mark = '✓' if i < len(EXPECTED_HEADERS) and h == EXPECTED_HEADERS[i] else '!'
            expected = EXPECTED_HEADERS[i] if i < len(EXPECTED_HEADERS) else '(unexpected)'
            print(f"  [{i:2}] {mark}  actual={h!r}  expected={expected!r}")
        print("\nKey columns:")
        for key, idx in COL.items():
            print(f"  {key:20} → [{idx}] {actual_headers[idx] if idx < len(actual_headers) else 'OUT OF RANGE'!r}")
        # Print first data row
        sample = next(ws.iter_rows(min_row=2, max_row=2), None)
        if sample:
            print("\nFirst data row values:")
            for i, cell in enumerate(sample):
                print(f"  [{i:2}] {actual_headers[i] if i < len(actual_headers) else '?'!r} = {cell.value!r}")
        return

    # ── Validate key column positions ─────────────────────────────────────────
    issues = []
    for key, idx in COL.items():
        actual = actual_headers[idx] if idx < len(actual_headers) else None
        if key == 'risk_status_enum':
            # Both col 3 and col 11 are named "Risk Status"; just verify it exists
            if actual != 'Risk Status':
                issues.append(f"  col {idx} ({key}): expected 'Risk Status', got {actual!r}")
        elif key == 'reviewed' and actual not in ('Reviewed', 'Analyst'):
            # Adjacent cols 9/10 — allow slight naming variation
            pass
        else:
            expected = EXPECTED_HEADERS[idx] if idx < len(EXPECTED_HEADERS) else None
            if actual != expected:
                issues.append(f"  col {idx} ({key}): expected {expected!r}, got {actual!r}")

    if issues:
        print("WARNING: Column layout may have changed. Run --print-columns to verify.")
        for issue in issues:
            print(issue)

    # ── Scope cutoff ──────────────────────────────────────────────────────────
    cutoff = datetime.now(tz=timezone.utc) - timedelta(days=scope_days)
    print(f"Scope: Added At >= {cutoff.date()} (last {scope_days} days)")

    # ── Pass 1 (ALL rows): build alert → URL mapping ──────────────────────────
    # XLSX cells carry only ONE hyperlink each. A cell with "Alert_A, Alert_B"
    # gives us one URL — we can't tell which part it belongs to from that cell alone.
    #
    # Strategy:
    #   single-alert cells  → unambiguous: alert_name → URL  (high confidence)
    #   multi-alert cells   → fallback only, used when no better source exists
    #
    # By scanning ALL rows (not just in-scope ones) we maximise the chance of
    # finding each alert in a single-alert cell somewhere in the sheet.

    alert_links: dict[str, str] = {}          # confirmed: single-alert cells
    alert_links_fallback: dict[str, str] = {} # uncertain: from multi-alert cells

    for row in ws.iter_rows(min_row=2):
        alert_cell = row[COL['alerts']]
        if not (alert_cell and alert_cell.hyperlink):
            continue
        url = alert_cell.hyperlink.target
        parts = [a.strip() for a in _cell_str(alert_cell).split(',') if a.strip()]
        if len(parts) == 1:
            alert_links[parts[0]] = url          # unambiguous
        else:
            for p in parts:
                if p not in alert_links and p not in alert_links_fallback:
                    alert_links_fallback[p] = url  # ambiguous fallback

    # Merge: confirmed wins
    combined_links: dict[str, str] = {**alert_links_fallback, **alert_links}

    # ── Pass 2: parse payment rows (apply scope filter) ───────────────────────
    rows = []
    all_alert_names: set[str] = set()
    skipped_scope = 0
    skipped_no_id = 0

    for row in ws.iter_rows(min_row=2):
        # Payment ID
        payment_id = _cell_str(row[COL['payment_id']])
        if not payment_id:
            skipped_no_id += 1
            continue

        # Added At — scope filter
        added_at_raw = row[COL['added_at']].value
        if added_at_raw is not None:
            if isinstance(added_at_raw, datetime):
                added_at = added_at_raw.replace(tzinfo=timezone.utc) \
                    if added_at_raw.tzinfo is None else added_at_raw
            else:
                try:
                    added_at = datetime.fromisoformat(str(added_at_raw)).replace(tzinfo=timezone.utc)
                except ValueError:
                    added_at = None
        else:
            added_at = None

        if added_at is None or added_at < cutoff:
            skipped_scope += 1
            continue

        alert_value = _cell_str(row[COL['alerts']])
        for a in alert_value.split(','):
            a = a.strip()
            if a:
                all_alert_names.add(a)

        rows.append({
            'payment_id':       payment_id,
            'alert_codes':      alert_value,
            'added_at':         added_at_raw,
            'reviewed':         _cell_str(row[COL['reviewed']]),
            'risk_status_enum': _cell_str(row[COL['risk_status_enum']]),
            'date_reviewed':    row[COL['date_reviewed']].value,
        })

    # ── Write outputs ─────────────────────────────────────────────────────────
    Path(out_dir).mkdir(parents=True, exist_ok=True)
    csv_path        = str(Path(out_dir) / 'sheet_data.csv')
    links_path      = str(Path(out_dir) / 'alert_links.json')
    names_path      = str(Path(out_dir) / 'unique_alerts.json')

    if not rows:
        print(f"WARNING: 0 rows passed the scope filter. Check Added At column format.")
    else:
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)

    with open(links_path, 'w', encoding='utf-8') as f:
        json.dump(combined_links, f, indent=2)

    # unique_alerts.json: sorted list of all unique alert names in scope
    # Used by SKILL.md Step 3 to drive per-alert Redash lookup
    with open(names_path, 'w', encoding='utf-8') as f:
        json.dump(sorted(all_alert_names), f, indent=2)

    no_url = sorted(all_alert_names - combined_links.keys())

    print(f"\nResults:")
    print(f"  Rows in scope (last {scope_days}d): {len(rows)}")
    print(f"  Skipped — no payment ID:  {skipped_no_id}")
    print(f"  Skipped — outside scope:  {skipped_scope}")
    print(f"  Unique alert names:       {len(all_alert_names)}")
    print(f"  Confirmed URL (single-cell source): {len(alert_links)}")
    print(f"  Fallback URL (multi-cell source):   {len(alert_links_fallback)}")
    if no_url:
        print(f"  NO URL found — need Redash name lookup: {no_url}")
    print(f"\nOutputs:\n  {csv_path}\n  {links_path}\n  {names_path}")


def main():
    parser = argparse.ArgumentParser(description='Parse Risk Intel Alerts XLSX')
    parser.add_argument('xlsx_path', help='Path to the downloaded XLSX file')
    parser.add_argument('--out-dir',     default='/tmp/risk_intel',
                        help='Output directory (default: /tmp/risk_intel)')
    parser.add_argument('--scope-days',  type=int, default=30,
                        help='Only include rows where Added At >= now - N days (default: 30; post-MVP: 90)')
    parser.add_argument('--print-columns', action='store_true',
                        help='Print detected columns and exit (use before first real run)')
    args = parser.parse_args()
    parse(args.xlsx_path, args.out_dir,
          scope_days=args.scope_days,
          print_columns=args.print_columns)


if __name__ == '__main__':
    main()
